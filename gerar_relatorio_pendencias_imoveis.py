"""
Relatório de Pendências – Conciliação Bens Imóveis (SIGGO × SISGEPAT)
======================================================================
Usa exatamente a mesma extração e regras do painel HTML
conciliacao_bens_imoveis_sisgepat.html (mesmos filtros INTIPOADM,
mesma Regra 7, mesmos totais).

Uso:
    python gerar_relatorio_pendencias_imoveis.py
    python gerar_relatorio_pendencias_imoveis.py --mes 7 --ano 2026
"""

import argparse, importlib.util, os, sys
from datetime import datetime
from pathlib import Path

import oracledb
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

_PASTA = Path(__file__).parent
load_dotenv(_PASTA / ".env")

ORACLE_USER = os.environ["ORACLE_USER"]
ORACLE_PASS = os.environ["ORACLE_PASS"]
ORACLE_DSN  = os.environ["ORACLE_DSN"]
IC = r"C:\oracle\instantclient_23_0"

# Carrega o script do painel HTML (contém extrair() com filtros corretos)
_HTML_SCRIPT = _PASTA / "extrair_conciliacao_bens_imoveis_sisgepat.py"
_spec = importlib.util.spec_from_file_location("painel_imoveis", str(_HTML_SCRIPT))
_painel = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_painel)
MESES = _painel.MESES
_core = _painel._core

# ── Paleta simples ────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

HDR      = _fill("1F497D")   # azul cabeçalho
HDR2     = _fill("4472C4")   # azul subeader / totais UG
TOT      = _fill("4472C4")   # azul total geral
ALT1     = _fill("FFFFFF")   # branco
ALT2     = _fill("DCE6F1")   # azul claro alternado
SUB      = _fill("BDD7EE")   # azul subtotal
RED_L    = _fill("F4CCCC")   # vermelho claro (A MENOR no SIGGO)
AMB_L    = _fill("FFE599")   # amarelo (A MAIOR no SIGGO)
GRN_L    = _fill("D9EAD3")   # verde claro
GRAY_L   = _fill("F2F2F2")   # cinza muito claro
WARN     = _fill("FCE4D6")   # laranja claro (metragem zero / alertas)
WHITE    = _fill("FFFFFF")

WF   = Font(name="Calibri", color="FFFFFF", bold=True, size=9)
HF   = Font(name="Calibri", color="FFFFFF", bold=True, size=9)
BF   = Font(name="Calibri", size=9)
BFB  = Font(name="Calibri", size=9, bold=True)
TF   = Font(name="Calibri", color="FFFFFF", bold=True, size=10)

_t = Side(style="thin",   color="B8CCE4")
_m = Side(style="medium", color="4472C4")
BT = Border(left=_t, right=_t, top=_t, bottom=_t)
BM = Border(left=_m, right=_m, top=_m, bottom=_m)

FMT  = '#,##0.00'
FMTI = '#,##0'
C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
L    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
R    = Alignment(horizontal="right",  vertical="center")


def _c(ws, row, col, value=None, *, fill=None, font=None, fmt=None,
       align=None, border=BT, bold=False, color=None, size=9):
    cell = ws.cell(row=row, column=col, value=value)
    f = font or Font(name="Calibri", size=size, bold=bold,
                     color=color if color else "000000")
    cell.font  = f
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt
    cell.alignment = align or L
    if border: cell.border = border
    return cell


def _hdr(ws, row, headers, fill=HDR, font=HF):
    for j, h in enumerate(headers, 1):
        _c(ws, row, j, h, fill=fill, font=font, align=C, border=BM)
    ws.row_dimensions[row].height = 28


def _titulo(ws, ncols, texto, sub_texto=None):
    """Faixa azul de título (linhas 1-2). Retorna próxima linha livre."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    _c(ws, 1, 1, texto, fill=HDR, font=TF, align=C, border=BM)
    ws.row_dimensions[1].height = 22
    r = 2
    if sub_texto:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        _c(ws, r, 1, sub_texto,
           fill=_fill("2F5496"),
           font=Font(name="Calibri", color="BDD7EE", size=8, italic=True),
           align=C, border=BT)
        ws.row_dimensions[r].height = 14
        r += 1
    return r


def _total(ws, row, ncols, label, valores, label_end=None):
    """Linha de total azul escuro."""
    end = label_end or ncols - len(valores)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end)
    _c(ws, row, 1, label, fill=TOT, font=WF, align=C, border=BM)
    for col, val in valores:
        _c(ws, row, col, val, fill=TOT, font=WF, fmt=FMT, align=R, border=BM)


def _subtotal(ws, row, ncols, label, valores, label_end=None):
    """Linha de subtotal azul médio."""
    end = label_end or ncols - len(valores)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end)
    _c(ws, row, 1, label, fill=SUB, font=BFB, align=C, border=BM)
    for col, val in valores:
        _c(ws, row, col, val, fill=SUB, font=BFB, fmt=FMT, align=R, border=BM)


def _width(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ─────────────────────────────────────────────────────────────────────────────
#  ABA 1 – PAINEL DE PENDÊNCIAS
# ─────────────────────────────────────────────────────────────────────────────
def _aba_painel(wb, quadro, eventos, transf, siggo_tomb, mes, ano):
    ws = wb.create_sheet("Painel de Pendências")
    ws.sheet_view.showGridLines = False

    ncols = 5
    sub = (f"Referência: {MESES[mes]}/{ano}  |  "
           "Emitido em: " + datetime.now().strftime("%d/%m/%Y %H:%M") +
           "  |  Administração Direta e Fundos do GDF (INTIPOADM = 1, 7)  |  "
           "Contas conciliáveis: 123210800/123210900/123211000/123219000/123219100")
    r = _titulo(ws, ncols, "CONCILIAÇÃO PATRIMONIAL SIGGO × SISGEPAT – BENS IMÓVEIS", sub)

    # KPIs da conciliação – idênticos ao painel HTML
    q_sis = quadro[quadro["CATEGORIA"] == "SISGEPAT"]
    tot_sis    = q_sis["SISGEPAT_VALOR"].sum()
    tot_sig    = q_sis["SIGGO_VALOR"].sum()
    dif_concil = q_sis["DIFERENCA"].sum()          # SISGEPAT − SIGGO, igual ao HTML
    n_ug_tot   = q_sis["UG"].nunique()
    n_ug_ok    = q_sis.groupby("UG")["DIFERENCA"].sum().abs().lt(0.01).sum()

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    _c(ws, r, 1, "SITUAÇÃO GERAL – contas conciliáveis (Adm. Direta e Fundos)",
       fill=HDR2, font=WF, align=C, border=BM)
    r += 1
    _hdr(ws, r, ["Indicador", "Valor (R$)", "Observação", "", ""], fill=_fill("375623"),
         font=Font(name="Calibri", color="FFFFFF", bold=True, size=9))
    r += 1

    def _kpi_row(label, val, obs, row, is_num=True, destaque=False):
        f = RED_L if (destaque and abs(val) > 0.01) else (GRN_L if destaque else ALT1)
        _c(ws, row, 1, label, fill=f, font=BFB, border=BT)
        if is_num and isinstance(val, float):
            _c(ws, row, 2, val, fill=f, fmt=FMT, align=R, border=BT,
               color="CC0000" if (destaque and abs(val) > 0.01) else "000000")
        else:
            _c(ws, row, 2, val, fill=f, fmt=FMTI if isinstance(val, int) else FMT,
               align=R, border=BT)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        _c(ws, row, 3, obs, fill=f, border=BT)

    _kpi_row("SISGEPAT (R$)",      float(tot_sis),    "Total inventariado no SISGEPAT", r)
    r += 1
    _kpi_row("SIGGO (R$)",         float(tot_sig),    "Total registrado no SIGGO (contas conciliáveis)", r)
    r += 1
    _kpi_row("Diferença (R$)",     float(dif_concil),
             "SISGEPAT − SIGGO; deve tender a zero à medida que as pendências são sanadas", r, destaque=True)
    r += 1
    _kpi_row("UGs conciliadas",    int(n_ug_ok),
             f"de {n_ug_tot} UGs ativas — conciliadas = diferença < R$ 0,01", r, is_num=False)
    r += 2

    # Pendências
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    _c(ws, r, 1, "RESUMO DE PENDÊNCIAS – providências a adotar pela unidade gestora",
       fill=HDR2, font=WF, align=C, border=BM)
    r += 1
    _hdr(ws, r, ["Tipo de Pendência", "Qtde", "Valor (R$)", "Aba de Detalhamento", "Providência Resumida"])
    r += 1

    # contadores
    n_transf  = len(transf) if transf is not None and len(transf) else 0
    vl_transf = float(transf["SIGGO_VALOR"].sum()) if n_transf else 0.0
    n_ev      = len(eventos) if eventos is not None and len(eventos) else 0
    vl_ev     = float(eventos["SIGGO_VALOR"].abs().sum()) if n_ev else 0.0

    ger = siggo_tomb.attrs.get("generico") if hasattr(siggo_tomb, "attrs") else None
    n_ig = vl_ig = 0
    if ger is not None and len(ger):
        ig_c = ger[ger["CONTA_CONTABIL"].isin(_core.CONTAS_SISGEPAT)]
        n_ig  = len(ig_c)
        vl_ig = float(ig_c["VALOR"].sum()) if "VALOR" in ig_c.columns else 0.0

    q_div = q_sis[q_sis["DIFERENCA"].abs() > 0.01]
    n_div  = len(q_div)
    vl_div = float(q_div["DIFERENCA"].abs().sum())

    linhas = [
        ("Transferências de imóveis sem evento no SIGGO", n_transf, vl_transf,
         "Transferências Pendentes",
         "Registrar evento de transferência no SIGGO (Direta→Direta: dispensa doação; Indireta→Direta: exige doação)"),
        ("Eventos contábeis pendentes – imóvel em conta errada", n_ev, vl_ev,
         "Eventos Pendentes",
         "Lançar evento de reclassificação de conta no SIGGO (art. 7.º do Dec. 16.109/94)"),
        ("Saldos em inscrição genérica (sem tombamento)", n_ig, vl_ig,
         "Inscrições Genéricas",
         "Reclassificar para tombamento individualizado (CI + número de tombamento)"),
        ("Divergências sem explicação identificada", n_div, vl_div,
         "Divergências a Investigar",
         "Investigar por UG e adotar a providência cabível (regularização, reclassificação ou contestação)"),
    ]
    fills = [ALT1, ALT2]
    for i, (desc, qtde, valor, aba, prov) in enumerate(linhas):
        f = fills[i % 2]
        _c(ws, r, 1, desc,  fill=f, font=BFB, border=BT)
        _c(ws, r, 2, qtde,  fill=f, fmt=FMTI, align=C, border=BT)
        _c(ws, r, 3, valor, fill=f, fmt=FMT,  align=R,  border=BT)
        _c(ws, r, 4, aba,   fill=f, align=C,  border=BT,
           color="1F497D", bold=True)
        _c(ws, r, 5, prov,  fill=f, border=BT)
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    _c(ws, r, 1,
       "Fonte e critérios: idênticos ao painel conciliacao_bens_imoveis_sisgepat.html — "
       "INTIPOADM ∈ {1, 7}, excluída UG 010101, Regra 7 corrigida (COD=00 Tipo=O sem terreno → Mobiliário Urbano). "
       "Diferença = SISGEPAT − SIGGO (negativo = SIGGO A MAIOR).",
       fill=GRAY_L, font=Font(name="Calibri", size=8, italic=True, color="595959"),
       align=L, border=Border())

    _width(ws, {"A": 46, "B": 9, "C": 22, "D": 26, "E": 50})


# ─────────────────────────────────────────────────────────────────────────────
#  ABA 2 – TRANSFERÊNCIAS PENDENTES
# ─────────────────────────────────────────────────────────────────────────────
def _aba_transferencias(wb, transf):
    ws = wb.create_sheet("Transferências Pendentes")
    ws.sheet_view.showGridLines = False
    ncols = 7
    sub = ("Imóveis registrados em unidades diferentes no SIGGO e no SISGEPAT. "
           "A UG de ORIGEM fica A MAIOR no SIGGO; a UG de DESTINO fica A MENOR. "
           "As diferenças se anulam no consolidado, mas cada unidade fica incorreta individualmente.")
    r = _titulo(ws, ncols, "TRANSFERÊNCIAS PENDENTES DE EVENTO NO SIGGO", sub)
    _hdr(ws, r, ["Tombamento", "UG Origem\n(SIGGO)", "Nome da UG de Origem",
                 "UG Destino\n(SISGEPAT)", "Nome da UG de Destino",
                 "Valor SIGGO (R$)", "Providência"])
    ws.freeze_panes = f"A{r + 1}"
    r += 1

    if transf is None or len(transf) == 0:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        _c(ws, r, 1, "Nenhuma transferência pendente identificada.", fill=GRN_L, align=C, border=BT)
        return

    fills = [ALT1, ALT2]
    vl = 0.0
    for i, row_data in enumerate(transf.itertuples(index=False)):
        tomb      = getattr(row_data, "TOMB_NORM", "")
        ug_orig   = getattr(row_data, "UG_SIGGO", "")
        nome_orig = getattr(row_data, "NOME_UG_SIGGO", "")
        ug_dest   = getattr(row_data, "UG_SISGEPAT", "")
        nome_dest = getattr(row_data, "NOME_UG_SISGEPAT", "")
        valor     = float(getattr(row_data, "SIGGO_VALOR", 0) or 0)
        prov      = getattr(row_data, "PROVIDENCIA", "")
        vl += valor
        f = fills[i % 2]
        _c(ws, r, 1, tomb,      fill=f, align=C, border=BT)
        _c(ws, r, 2, ug_orig,   fill=f, align=C, border=BT)
        _c(ws, r, 3, nome_orig, fill=f, border=BT)
        _c(ws, r, 4, ug_dest,   fill=f, align=C, border=BT)
        _c(ws, r, 5, nome_dest, fill=f, border=BT)
        _c(ws, r, 6, valor,     fill=f, fmt=FMT, align=R, border=BT)
        _c(ws, r, 7, prov,      fill=f, border=BT)
        ws.row_dimensions[r].height = 30
        r += 1

    _total(ws, r, ncols, f"TOTAL – {len(transf)} imóvel(is)", [(6, vl)], label_end=5)
    _width(ws, {"A": 14, "B": 10, "C": 38, "D": 10, "E": 38, "F": 20, "G": 55})


# ─────────────────────────────────────────────────────────────────────────────
#  ABA 3 – EVENTOS CONTÁBEIS PENDENTES
# ─────────────────────────────────────────────────────────────────────────────
def _aba_eventos(wb, eventos):
    ws = wb.create_sheet("Eventos Pendentes")
    ws.sheet_view.showGridLines = False
    ncols = 8
    sub = ("Imóveis classificados em conta diferente nos dois sistemas. "
           "A unidade gestora deve lançar o evento contábil de reclassificação no SIGGO, "
           "instruído com a documentação do art. 7.º do Decreto 16.109/94.")
    r = _titulo(ws, ncols, "EVENTOS CONTÁBEIS PENDENTES NO SIGGO – por imóvel (tombamento)", sub)
    _hdr(ws, r, ["UG", "Unidade Gestora", "Tombamento",
                 "Conta Atual (SIGGO)", "Conta Devida (SISGEPAT)",
                 "Valor SIGGO (R$)", "Valor SISGEPAT (R$)", "Providência"])
    ws.freeze_panes = f"A{r + 1}"
    r += 1

    if eventos is None or len(eventos) == 0:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        _c(ws, r, 1, "Nenhum evento contábil pendente identificado.", fill=GRN_L, align=C, border=BT)
        return

    fills = [ALT1, ALT2]
    tot_sig = tot_sis = 0.0
    n_tot = 0

    for ug, grp in eventos.groupby("UG", sort=False):
        nome_ug = grp["NOME_UG"].iloc[0] if "NOME_UG" in grp.columns else ""
        for i, rd in enumerate(grp.itertuples(index=False)):
            tomb  = getattr(rd, "TOMB_NORM", "")
            c_sig = getattr(rd, "CONTA_SIGGO", "")
            c_sis = getattr(rd, "CONTA_SISGEPAT", "")
            v_sig = float(getattr(rd, "SIGGO_VALOR", 0) or 0)
            v_sis = float(getattr(rd, "SISGEPAT_VALOR", 0) or 0)
            prov  = getattr(rd, "EVENTO", "")
            f = fills[n_tot % 2]
            _c(ws, r, 1, ug,      fill=f, align=C, border=BT)
            _c(ws, r, 2, nome_ug, fill=f, border=BT)
            _c(ws, r, 3, tomb,    fill=f, align=C, border=BT)
            _c(ws, r, 4, c_sig,   fill=f, border=BT)
            _c(ws, r, 5, c_sis,   fill=f, border=BT)
            _c(ws, r, 6, v_sig,   fill=f, fmt=FMT, align=R, border=BT)
            _c(ws, r, 7, v_sis,   fill=f, fmt=FMT, align=R, border=BT)
            _c(ws, r, 8, prov,    fill=f, border=BT)
            ws.row_dimensions[r].height = 30
            tot_sig += v_sig; tot_sis += v_sis; n_tot += 1; r += 1

        sub_sig = float(grp["SIGGO_VALOR"].sum())
        sub_sis = float(grp["SISGEPAT_VALOR"].sum())
        _subtotal(ws, r, ncols, f"Subtotal {ug} – {len(grp)} imóvel(is)",
                  [(6, sub_sig), (7, sub_sis)], label_end=5)
        r += 1

    _total(ws, r, ncols, f"TOTAL – {n_tot} imóvel(is)",
           [(6, tot_sig), (7, tot_sis)], label_end=5)
    _width(ws, {"A": 10, "B": 38, "C": 14, "D": 35, "E": 35, "F": 20, "G": 20, "H": 55})


# ─────────────────────────────────────────────────────────────────────────────
#  ABA 4 – INSCRIÇÕES GENÉRICAS
# ─────────────────────────────────────────────────────────────────────────────
def _aba_inscricoes_genericas(wb, siggo_tomb):
    ws = wb.create_sheet("Inscrições Genéricas")
    ws.sheet_view.showGridLines = False
    ncols = 6
    sub = ("Saldos do SIGGO em inscrição genérica (ex.: IM9999999) sem tombamento individualizado. "
           "Enquanto não reclassificados, a conciliação imóvel a imóvel é impossível. "
           "Prioridade: linhas marcadas em laranja (contas conciliáveis com o SISGEPAT).")
    r = _titulo(ws, ncols, "SALDOS DO SIGGO EM INSCRIÇÃO GENÉRICA – sem tombamento individualizado", sub)
    _hdr(ws, r, ["UG", "Unidade Gestora", "Conta Contábil", "Descrição da Conta",
                 "Conciliável com SISGEPAT?", "Valor (R$)"])
    ws.freeze_panes = f"A{r + 1}"
    r += 1

    ger = siggo_tomb.attrs.get("generico") if hasattr(siggo_tomb, "attrs") else None
    if ger is None or len(ger) == 0:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        _c(ws, r, 1, "Nenhuma inscrição genérica identificada.", fill=GRN_L, align=C, border=BT)
        return

    if "CONCILIAVEL" not in ger.columns:
        ger = ger.copy()
        ger["CONCILIAVEL"] = ger["CONTA_CONTABIL"].isin(_core.CONTAS_SISGEPAT).map({True: "SIM", False: "NÃO"})

    ger_sorted = ger.sort_values(["CONCILIAVEL", "COD_UG", "VALOR"], ascending=[False, True, False])

    fills = [ALT1, ALT2]
    vl_c = vl_nc = 0.0
    for i, rd in enumerate(ger_sorted.itertuples(index=False)):
        ug    = str(getattr(rd, "COD_UG", "")).zfill(6)
        nome  = getattr(rd, "NOME_UG", "")
        conta = getattr(rd, "CONTA_CONTABIL", "")
        desc  = _core.DESCRICAO_CONTA.get(str(conta), "")
        conc  = getattr(rd, "CONCILIAVEL", "")
        valor = float(getattr(rd, "VALOR", 0) or 0)
        f     = WARN if conc == "SIM" else fills[i % 2]
        _c(ws, r, 1, ug,    fill=f, align=C, border=BT)
        _c(ws, r, 2, nome,  fill=f, border=BT)
        _c(ws, r, 3, conta, fill=f, align=C, border=BT)
        _c(ws, r, 4, desc,  fill=f, border=BT)
        _c(ws, r, 5, conc,  fill=f, align=C, border=BT,
           bold=(conc == "SIM"), color="C00000" if conc == "SIM" else "000000")
        _c(ws, r, 6, valor, fill=f, fmt=FMT, align=R, border=BT)
        if conc == "SIM": vl_c += valor
        else:             vl_nc += valor
        r += 1

    _total(ws, r, ncols,
           f"TOTAL em inscrições genéricas CONCILIÁVEIS (ação prioritária) – {len(ger[ger['CONCILIAVEL']=='SIM'])} registros",
           [(6, vl_c)], label_end=5)
    _width(ws, {"A": 10, "B": 42, "C": 14, "D": 38, "E": 22, "F": 20})


# ─────────────────────────────────────────────────────────────────────────────
#  ABA 5 – DIVERGÊNCIAS A INVESTIGAR
# ─────────────────────────────────────────────────────────────────────────────
def _aba_divergencias(wb, quadro):
    ws = wb.create_sheet("Divergências a Investigar")
    ws.sheet_view.showGridLines = False
    ncols = 7
    sub = ("Combinações UG + conta com diferença entre SIGGO e SISGEPAT, ordenadas pelo valor absoluto. "
           "Diferença = SISGEPAT − SIGGO  |  Negativo = SIGGO A MAIOR  |  Positivo = SISGEPAT A MAIOR.")
    r = _titulo(ws, ncols, "DIVERGÊNCIAS A INVESTIGAR – Adm. Direta e Fundos (contas conciliáveis)", sub)
    _hdr(ws, r, ["UG", "Unidade Gestora", "Conta", "Descrição da Conta",
                 "SISGEPAT (R$)", "SIGGO (R$)", "Diferença (R$)"])
    ws.freeze_panes = f"A{r + 1}"
    r += 1

    q_div = (quadro[(quadro["CATEGORIA"] == "SISGEPAT") & (quadro["DIFERENCA"].abs() > 0.01)]
             .copy()
             .sort_values("DIFERENCA", key=lambda s: s.abs(), ascending=False))

    if len(q_div) == 0:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        _c(ws, r, 1, "Nenhuma divergência identificada.", fill=GRN_L, align=C, border=BT)
        return

    fills = [ALT1, ALT2]
    for i, rd in enumerate(q_div.itertuples(index=False)):
        ug   = getattr(rd, "UG", "")
        nome = getattr(rd, "NOME_UG", "")
        conta= getattr(rd, "CONTA", "")
        desc = getattr(rd, "DESCRICAO_CONTA", "")
        sis  = float(getattr(rd, "SISGEPAT_VALOR", 0) or 0)
        sig  = float(getattr(rd, "SIGGO_VALOR",  0) or 0)
        dif  = float(getattr(rd, "DIFERENCA",    0) or 0)
        f_dif = RED_L if dif < -0.01 else (AMB_L if dif > 0.01 else fills[i % 2])
        f     = fills[i % 2]
        _c(ws, r, 1, ug,   fill=f, align=C, border=BT)
        _c(ws, r, 2, nome, fill=f, border=BT)
        _c(ws, r, 3, conta,fill=f, align=C, border=BT)
        _c(ws, r, 4, desc, fill=f, border=BT)
        _c(ws, r, 5, sis,  fill=f, fmt=FMT, align=R, border=BT)
        _c(ws, r, 6, sig,  fill=f, fmt=FMT, align=R, border=BT)
        _c(ws, r, 7, dif,  fill=f_dif, fmt=FMT, align=R, border=BT, bold=True,
           color="CC0000" if dif < -0.01 else ("7F6000" if dif > 0.01 else "000000"))
        r += 1

    _total(ws, r, ncols,
           f"TOTAL – {len(q_div)} combinação(ões) com divergência",
           [(5, float(q_div["SISGEPAT_VALOR"].sum())),
            (6, float(q_div["SIGGO_VALOR"].sum())),
            (7, float(q_div["DIFERENCA"].sum()))],
           label_end=4)
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    _c(ws, r, 1,
       "Cores – Diferença:  🔴 vermelho = SIGGO A MAIOR que SISGEPAT (diferença negativa)  |  "
       "🟡 amarelo = SISGEPAT A MAIOR que SIGGO (diferença positiva)",
       fill=GRAY_L, font=Font(name="Calibri", size=8, italic=True, color="595959"),
       align=L, border=Border())
    _width(ws, {"A": 10, "B": 42, "C": 12, "D": 35, "E": 20, "F": 20, "G": 20})


# ─────────────────────────────────────────────────────────────────────────────
#  ABA 6 – METRAGEM ZERO (SISGEPAT)
# ─────────────────────────────────────────────────────────────────────────────
def _aba_metragem_zero(wb):
    ws = wb.create_sheet("Metragem Zero")
    ws.sheet_view.showGridLines = False
    ncols = 7
    sub = ("Terrenos e edificações no SISGEPAT com campo METRAGEM nulo, vazio ou igual a zero. "
           "Podem indicar omissão de atualização cadastral ou registros incompletos.")
    ncols = 7
    r = _titulo(ws, ncols, "SISGEPAT – IMÓVEIS COM METRAGEM ZERO OU NÃO INFORMADA", sub)
    _hdr(ws, r, ["Tombamento", "Tipo", "Unidade / Localização (SISGEPAT)",
                 "Metragem", "Valor (R$)", "Processo", "SR"])
    ws.freeze_panes = f"A{r + 1}"
    r += 1

    SQL = r"""
        SELECT t.TOMBAMENTO, 'TERRENO' AS TIPO,
               t.LOCALIZACAO AS UG_NOME,
               t.PROCESSO,
               t.METRAGEM,
               t.VALOR,
               t.SR_CODIGO
        FROM SIGGO.PAT_TERRENOS t
        WHERE (t.METRAGEM IS NULL
               OR TRIM(t.METRAGEM) = ''
               OR REGEXP_LIKE(TRIM(t.METRAGEM), '^0+([,.]0*)?$'))
        UNION ALL
        SELECT e.TOMBAMENTO, 'EDIFICACAO' AS TIPO,
               e.LOCALIZACAO AS UG_NOME,
               e.PROCESSO,
               e.METRAGEM,
               e.VALOR,
               e.SR_CODIGO
        FROM SIGGO.PAT_EDIFICACAO e
        WHERE (e.METRAGEM IS NULL
               OR TRIM(e.METRAGEM) = ''
               OR REGEXP_LIKE(TRIM(e.METRAGEM), '^0+([,.]0*)?$'))
        ORDER BY TIPO, TOMBAMENTO
    """

    try:
        oracledb.init_oracle_client(lib_dir=IC)
    except Exception:
        pass

    with oracledb.connect(user=ORACLE_USER, password=ORACLE_PASS, dsn=ORACLE_DSN) as conn:
        df = pd.read_sql(SQL, conn)

    if df.empty:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        _c(ws, r, 1, "Nenhum imóvel com metragem zero identificado.", fill=GRN_L, align=C, border=BT)
        return

    fills = [ALT1, ALT2]
    for i, (_, rd) in enumerate(df.iterrows()):
        metr  = rd.get("METRAGEM", "")
        valor = rd.get("VALOR", None)
        sr    = rd.get("SR_CODIGO", "")
        f = fills[i % 2]
        _c(ws, r, 1, rd.get("TOMBAMENTO", ""), fill=f, align=C, border=BT)
        _c(ws, r, 2, rd.get("TIPO",       ""), fill=f, align=C, border=BT)
        _c(ws, r, 3, rd.get("UG_NOME",    ""), fill=f, border=BT)
        _c(ws, r, 4, str(metr) if metr is not None else "(nulo)", fill=f, align=C, border=BT,
           color="C00000")
        if valor is not None:
            _c(ws, r, 5, float(valor), fill=f, fmt=FMT, align=R, border=BT)
        else:
            _c(ws, r, 5, None, fill=f, border=BT)
        _c(ws, r, 6, rd.get("PROCESSO", ""), fill=f, border=BT)
        _c(ws, r, 7, str(sr), fill=f, align=C, border=BT)
        r += 1

    n_terr = len(df[df["TIPO"] == "TERRENO"])
    n_edif = len(df[df["TIPO"] == "EDIFICACAO"])
    _total(ws, r, ncols,
           f"TOTAL – {len(df)} registro(s): {n_terr} terreno(s), {n_edif} edificação(ões)",
           [(5, float(df["VALOR"].sum()))], label_end=4)
    _width(ws, {"A": 14, "B": 14, "C": 45, "D": 10, "E": 20, "F": 28, "G": 6})


# ─────────────────────────────────────────────────────────────────────────────
#  ABA 7 – ESTUDOS E PROJETOS SEM CONTINUIDADE
# ─────────────────────────────────────────────────────────────────────────────
def _aba_estudos_sem_continuidade(wb):
    ws = wb.create_sheet("Estudos Sem Continuidade")
    ws.sheet_view.showGridLines = False
    ncols = 8
    sub = ("Estudos e projetos (ER=91, EL=51, SD=01) cujo tombamento NÃO possui nenhum outro registro "
           "com SD=02 (em obra) ou SD=10 (concluída). Indica estudos que não evoluíram para execução real de obra.")
    ncols = 9
    r = _titulo(ws, ncols, "ESTUDOS E PROJETOS SEM CONTINUIDADE – SISGEPAT (PAT_OBRAS)", sub)
    _hdr(ws, r, ["Tombamento", "UG", "Nome da Unidade Gestora", "Cód. Obra", "Descrição da Obra",
                 "Valor (R$)", "Data Medição", "Processo", "Cód. Localização"])
    ws.freeze_panes = f"A{r + 1}"
    r += 1

    SQL = r"""
        SELECT
            o.IM_TOMBAMENTO   AS TOMBAMENTO,
            o.UG_CODIGO       AS UG,
            NVL(pi.LOCALIZACAO, o.UG_CODIGO) AS UG_NOME,
            o.OB_CODIGO       AS COD_OBRA,
            o.OB_DESCRICAO    AS DESCRICAO,
            o.OB_VALOR        AS VALOR,
            o.OB_DTMEDICAO    AS DT_MEDICAO,
            o.OB_PROCESSO     AS PROCESSO,
            o.LO_CODIGO       AS LO_CODIGO
        FROM SIGGO.PAT_OBRAS o
        LEFT JOIN SIGGO.PAT_IMOVEIS pi ON pi.TOMBAMENTO = o.IM_TOMBAMENTO
        WHERE o.ER_CODIGO = '91'
          AND o.EL_CODIGO = '51'
          AND o.SD_CODIGO = '01'
          AND o.IM_TOMBAMENTO NOT IN (
              SELECT DISTINCT IM_TOMBAMENTO
              FROM SIGGO.PAT_OBRAS
              WHERE ER_CODIGO = '91'
                AND SD_CODIGO IN ('02', '10')
                AND IM_TOMBAMENTO IS NOT NULL
          )
        ORDER BY o.OB_VALOR DESC NULLS LAST, o.IM_TOMBAMENTO
    """

    try:
        oracledb.init_oracle_client(lib_dir=IC)
    except Exception:
        pass

    with oracledb.connect(user=ORACLE_USER, password=ORACLE_PASS, dsn=ORACLE_DSN) as conn:
        df = pd.read_sql(SQL, conn)

    if df.empty:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        _c(ws, r, 1, "Nenhum estudo sem continuidade encontrado.", fill=GRN_L, align=C, border=BT)
        return

    fills = [ALT1, ALT2]
    for i, rd in df.iterrows():
        valor = rd.get("VALOR", None)
        f = fills[i % 2]
        _c(ws, r, 1, rd.get("TOMBAMENTO", ""), fill=f, align=C, border=BT)
        _c(ws, r, 2, rd.get("UG",         ""), fill=f, align=C, border=BT)
        _c(ws, r, 3, rd.get("UG_NOME",    ""), fill=f, border=BT)
        _c(ws, r, 4, rd.get("COD_OBRA",   ""), fill=f, align=C, border=BT)
        _c(ws, r, 5, rd.get("DESCRICAO",  ""), fill=f, border=BT)
        if valor is not None:
            _c(ws, r, 6, float(valor), fill=f, fmt=FMT, align=R, border=BT)
        else:
            _c(ws, r, 6, None, fill=f, border=BT)
        _c(ws, r, 7, rd.get("DT_MEDICAO", ""), fill=f, align=C, border=BT)
        _c(ws, r, 8, rd.get("PROCESSO",   ""), fill=f, border=BT)
        _c(ws, r, 9, rd.get("LO_CODIGO",  ""), fill=f, align=C, border=BT)
        r += 1

    vl_total = float(df["VALOR"].sum()) if "VALOR" in df.columns else 0.0
    _total(ws, r, ncols,
           f"TOTAL – {len(df)} estudo(s)/projeto(s) sem continuidade",
           [(6, vl_total)], label_end=5)

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    _c(ws, r, 1,
       "Critério: estudos com ER=91 (Obras em Andamento), EL=51, SD=01 (Estudo/Projeto) "
       "cujo tombamento NÃO possui nenhum outro registro com SD=02 (Em Obra) ou SD=10 (Concluída). "
       "Estudos que evoluíram para obra real são excluídos desta listagem.",
       fill=GRAY_L, font=Font(name="Calibri", size=8, italic=True, color="595959"),
       align=L, border=Border())

    _width(ws, {"A": 14, "B": 10, "C": 38, "D": 12, "E": 38, "F": 20, "G": 14, "H": 22, "I": 16})


# ─────────────────────────────────────────────────────────────────────────────
#  GERAÇÃO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────
def gerar(mes, ano):
    print(f"\n{'=' * 60}")
    print(f"  RELATÓRIO DE PENDÊNCIAS – BENS IMÓVEIS  {MESES[mes]}/{ano}")
    print(f"{'=' * 60}\n")

    print("[1/3] Extraindo dados do Oracle (mesma fonte do painel HTML)…")
    quadro, eventos, transf, achados, siggo_tomb = _painel.extrair(mes, ano)
    print(f"  Quadro: {len(quadro)} linhas  |  Eventos: {len(eventos) if eventos is not None else 0}"
          f"  |  Transferências: {len(transf) if transf is not None else 0}")

    print("[2/3] Montando relatório Excel…")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _aba_painel(wb, quadro, eventos, transf, siggo_tomb, mes, ano)
    _aba_transferencias(wb, transf)
    _aba_eventos(wb, eventos)
    _aba_inscricoes_genericas(wb, siggo_tomb)
    _aba_divergencias(wb, quadro)
    _aba_metragem_zero(wb)
    _aba_estudos_sem_continuidade(wb)

    print("[3/3] Salvando…")
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome = f"Relatorio_Pendencias_Imoveis_{ano}_{mes:02d}_{ts}.xlsx"
    saida = _PASTA / nome
    wb.save(str(saida))
    print(f"\n  OK! Arquivo salvo: {saida}\n")
    return saida


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--mes", type=int, default=None)
    p.add_argument("--ano", type=int, default=None)
    a = p.parse_args()
    hoje = datetime.today()
    mes  = a.mes or (hoje.month - 1 or 12)
    ano  = a.ano or (hoje.year if hoje.month > 1 else hoje.year - 1)
    gerar(mes, ano)


if __name__ == "__main__":
    main()
